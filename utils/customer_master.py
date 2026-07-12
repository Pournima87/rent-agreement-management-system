import os

from openpyxl import Workbook, load_workbook


MASTER_FILE = "records/customer_master.xlsx"


# =====================================
# Create Master Database
# =====================================

def create_master_file():

    os.makedirs("records", exist_ok=True)

    if os.path.exists(MASTER_FILE):
        return

    wb = Workbook()

    ws = wb.active

    ws.title = "Customer Master"

    ws.append([

        # -----------------------------
        # CUSTOMER
        # -----------------------------

        "Customer ID",
        "Created Date",

        # -----------------------------
        # OWNER
        # -----------------------------

        "Owner Name",
        "Owner Age",
        "Owner Address",
        "Owner Mobile",

        # -----------------------------
        # TENANT
        # -----------------------------

        "Tenant Name",
        "Tenant Age",
        "Tenant Address",
        "Tenant Mobile",

        "Aadhaar",
        "PAN",

        # -----------------------------
        # PROPERTY
        # -----------------------------

        "Property Type",
        "Property Address",
        "Purpose",

        # -----------------------------
        # AGREEMENT
        # -----------------------------

        "Agreement Date",

        "Start Date",

        "End Date",

        "Agreement Months",

        "Rent",

        "Rent Words",

        "Deposit",

        "Deposit Words",

        "Rent Pay Date",

        # -----------------------------
        # FAMILY
        # -----------------------------

        "Family1 Name",
        "Family1 Relation",

        "Family2 Name",
        "Family2 Relation",

        "Family3 Name",
        "Family3 Relation",

        "Family4 Name",
        "Family4 Relation",

        "Family5 Name",
        "Family5 Relation",

        "Family6 Name",
        "Family6 Relation",

        "Family7 Name",
        "Family7 Relation",

        # -----------------------------
        # POLICE NOC
        # -----------------------------

        "Previous Address",

        "Male Count",

        "Female Count",

        "Child Count",

        "Work Type",

        "Office Details",

        "Reference 1",

        "Reference 2",

        "Agent Details",

        # -----------------------------
        # DOCUMENTS
        # -----------------------------

        "Agreement PDF",

        "Police NOC PDF",

        # -----------------------------
        # SYSTEM
        # -----------------------------

        "Status",

        "Reminder Sent",

        "Last Reminder Date"

    ])

    wb.save(MASTER_FILE)


# =====================================
# Save Full Customer
# =====================================

def save_customer_master(data):

    create_master_file()

    wb = load_workbook(MASTER_FILE)

    ws = wb.active

    ws.append([

        data.get("customer_id", ""),
        data.get("created_date", ""),

        data.get("owner_name", ""),
        data.get("owner_age", ""),
        data.get("owner_address", ""),
        data.get("owner_mobile", ""),

        data.get("tenant_name", ""),
        data.get("tenant_age", ""),
        data.get("tenant_address", ""),
        data.get("tenant_mobile", ""),

        data.get("aadhaar", ""),
        data.get("pan", ""),

        data.get("property_type", ""),
        data.get("property_address", ""),
        data.get("purpose", ""),

        data.get("agreement_date", ""),

        data.get("start_date", ""),

        data.get("end_date", ""),

        data.get("agreement_months", ""),

        data.get("rent", ""),
        data.get("rent_words", ""),

        data.get("deposit", ""),
        data.get("deposit_words", ""),

        data.get("rent_pay_date", ""),

        data.get("family_1_name", ""),
        data.get("family_1_relation", ""),

        data.get("family_2_name", ""),
        data.get("family_2_relation", ""),

        data.get("family_3_name", ""),
        data.get("family_3_relation", ""),

        data.get("family_4_name", ""),
        data.get("family_4_relation", ""),

        data.get("family_5_name", ""),
        data.get("family_5_relation", ""),

        data.get("family_6_name", ""),
        data.get("family_6_relation", ""),

        data.get("family_7_name", ""),
        data.get("family_7_relation", ""),

        data.get("previous_address", ""),

        data.get("male_count", ""),

        data.get("female_count", ""),

        data.get("child_count", ""),

        data.get("work_type", ""),

        data.get("office_details", ""),

        data.get("reference_1", ""),

        data.get("reference_2", ""),

        data.get("agent_details", ""),

        data.get("agreement_pdf", ""),

        data.get("noc_pdf", ""),

        "ACTIVE",

        "NO",

        ""

    ])

    wb.save(MASTER_FILE)

# =====================================
# Get Customer By ID
# =====================================

def get_customer_by_id(customer_id):

    create_master_file()

    wb = load_workbook(MASTER_FILE)

    ws = wb.active

    headers = []

    for cell in ws[1]:
        headers.append(cell.value)

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0] == customer_id:

            customer = dict(zip(headers, row))

            return {

                "customer_id": customer.get("Customer ID", ""),
                "created_date": customer.get("Created Date", ""),

                "owner_name": customer.get("Owner Name", ""),
                "owner_age": customer.get("Owner Age", ""),
                "owner_address": customer.get("Owner Address", ""),
                "owner_mobile": customer.get("Owner Mobile", ""),

                "tenant_name": customer.get("Tenant Name", ""),
                "tenant_age": customer.get("Tenant Age", ""),
                "tenant_address": customer.get("Tenant Address", ""),
                "tenant_mobile": customer.get("Tenant Mobile", ""),

                "aadhaar": customer.get("Aadhaar", ""),
                "pan": customer.get("PAN", ""),

                "property_type": customer.get("Property Type", ""),
                "property_address": customer.get("Property Address", ""),
                "purpose": customer.get("Purpose", ""),

                "agreement_date": customer.get("Agreement Date", ""),
                "start_date": customer.get("Start Date", ""),
                "end_date": customer.get("End Date", ""),
                "agreement_months": customer.get("Agreement Months", ""),

                "rent": customer.get("Rent", ""),
                "rent_words": customer.get("Rent Words", ""),

                "deposit": customer.get("Deposit", ""),
                "deposit_words": customer.get("Deposit Words", ""),

                "rent_pay_date": customer.get("Rent Pay Date", ""),

                # -----------------------------
                # FAMILY
                # -----------------------------

                "family_1_name": customer.get("Family1 Name", ""),
                "family_1_relation": customer.get("Family1 Relation", ""),

                "family_2_name": customer.get("Family2 Name", ""),
                "family_2_relation": customer.get("Family2 Relation", ""),

                "family_3_name": customer.get("Family3 Name", ""),
                "family_3_relation": customer.get("Family3 Relation", ""),

                "family_4_name": customer.get("Family4 Name", ""),
                "family_4_relation": customer.get("Family4 Relation", ""),

                "family_5_name": customer.get("Family5 Name", ""),
                "family_5_relation": customer.get("Family5 Relation", ""),

                "family_6_name": customer.get("Family6 Name", ""),
                "family_6_relation": customer.get("Family6 Relation", ""),

                "family_7_name": customer.get("Family7 Name", ""),
                "family_7_relation": customer.get("Family7 Relation", ""),

                # -----------------------------
                # POLICE NOC
                # -----------------------------

                "previous_address": customer.get("Previous Address", ""),

                "male_count": customer.get("Male Count", ""),
                "female_count": customer.get("Female Count", ""),
                "child_count": customer.get("Child Count", ""),

                "work_type": customer.get("Work Type", ""),

                "office_details": customer.get("Office Details", ""),

                "reference_1": customer.get("Reference 1", ""),
                "reference_2": customer.get("Reference 2", ""),

                "agent_details": customer.get("Agent Details", "")

            }

    return None